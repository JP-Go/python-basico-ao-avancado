import openpyxl as xl


if __name__ == "__main__":
    relative_file_path = "./public/pedidos.xlsx"
    # xl.load_workbook(filepath): Abre o pasta de planilhas no caminho filepath
    workbook = xl.load_workbook(relative_file_path)
    # xl.sheetnames: Nome das planilhas no pasta de planilhas
    sheets = workbook.sheetnames
    print(sheets)

    # Planilhas da pasta são indexadas como chaves em dicionários (por strings)
    orders = workbook["Página1"]

    # Células são selecionadas como chaves em dicionários (retorna um objeto Cell)
    # print(orders["B4"])
    # print(orders["B4"].value)

    # Pode-se selecionar uma coluna ou linha inteira como descrito a seguir
    # Retorna uma tupla de valores
    # b_column = orders["B"]
    # for cell in b_column:
    #     print(cell.value)

    # Pode-se selecionar um intervalo de células da seguinte forma
    # retorna uma tupla de tuplas
    # arange = orders["a1:c2"]
    # As linhas e colunas são iteradas na ordem de linha-coluna
    # for line in arange:
    #     # ordem de leitura
    #     #   1 2 3
    #     #   4 5 6
    #     for col in line:
    #         print(col.value)

    # Pode-se, por fim, iterar sobre a planilha toda:
    # for line in orders:
    #     for col in line:
    #         print(col.value)

    for line in orders:
        for i in range(len(line)):
            if line[i].value is not None:
                sep = " " if i < len(line) else "\n"
                print(sep)
                print(line[i].value, sep=sep)
